"""Импорт таблицы КБЖУ сети в карточки блюд.

Запуск: uv run python -m app.import_nutrition
Названия в таблице и в меню совпадают не буква в букву, поэтому сопоставляем
по упрощённому виду: без регистра, без лишних пробелов и знаков.
"""

import asyncio
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.config import settings
from app.core.db import SessionLocal, engine
from app.models.menu import Dish

SOURCE = Path(__file__).resolve().parents[2] / "КБЖУ 2024.xlsx"
HEADER_ROW = 7
TENANT_ID = settings.default_tenant_id

# Колонки: название, выход в кг, затем на 100 г, затем на порцию
COL_NAME = 1
COL_WEIGHT = 2
COL_PORTION_FATS = 7
COL_PORTION_PROTEINS = 9
COL_PORTION_CARBS = 10
COL_PORTION_CALORIES = 11

NOISE_RE = re.compile(r"[^а-яёa-z0-9 ]")

# В таблице блюда часто названы с типом впереди, в меню — без него
LEADING_TYPE_RE = re.compile(
    r"^(пицца|салат|суп|паста|десерт|напиток|закуска|ризотто|соус|горячее)\s+"
)


@dataclass(slots=True)
class Nutrition:
    name: str
    weight_grams: int | None
    proteins_g: float | None
    fats_g: float | None
    carbs_g: float | None
    calories: int | None


def simplify(name: str) -> str:
    """«Пицца Поло Песто» и «Поло песто» должны совпасть."""
    text = name.lower().replace("ё", "е").replace("«", " ").replace("»", " ")
    text = NOISE_RE.sub(" ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return LEADING_TYPE_RE.sub("", text).strip() or text


def to_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)

    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def read_table() -> list[Nutrition]:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[Nutrition] = []
    for row in list(sheet.iter_rows(values_only=True))[HEADER_ROW:]:
        raw_name = row[COL_NAME]
        if not raw_name or not str(raw_name).strip():
            continue

        weight_kg = to_number(row[COL_WEIGHT])
        calories = to_number(row[COL_PORTION_CALORIES])

        rows.append(
            Nutrition(
                name=str(raw_name).strip(),
                weight_grams=round(weight_kg * 1000) if weight_kg else None,
                proteins_g=to_number(row[COL_PORTION_PROTEINS]),
                fats_g=to_number(row[COL_PORTION_FATS]),
                carbs_g=to_number(row[COL_PORTION_CARBS]),
                calories=round(calories) if calories else None,
            )
        )
    return rows


async def run() -> None:
    table = read_table()
    print(f"Строк в таблице: {len(table)}")

    by_key = {simplify(item.name): item for item in table}
    keys = list(by_key)
    matched = fuzzy = missed = 0

    def find(name: str) -> tuple[Nutrition | None, bool]:
        """Точное совпадение, затем вхождение, затем похожесть."""
        key = simplify(name)
        exact = by_key.get(key)
        if exact is not None:
            return exact, False

        # Вхождение принимаем только когда названия сопоставимы по длине:
        # иначе «Тальятелле с грибами» цепляет данные блюда «с грибами и курицей»
        contained = [
            other
            for other in keys
            if len(other) >= 10
            and len(key) >= 10
            and (other in key or key in other)
            and min(len(other), len(key)) / max(len(other), len(key)) >= 0.72
        ]
        if contained:
            return by_key[max(contained, key=len)], True

        best, score = None, 0.0
        for other in keys:
            ratio = SequenceMatcher(None, key, other).ratio()
            if ratio > score:
                best, score = other, ratio
        if best is not None and score >= 0.87:
            return by_key[best], True

        return None, False

    async with SessionLocal() as session:
        dishes = (await session.scalars(select(Dish).where(Dish.tenant_id == TENANT_ID))).all()

        for dish in dishes:
            item, approximate = find(dish.name)
            if item is None:
                missed += 1
                continue
            if approximate:
                fuzzy += 1

            dish.proteins_g = item.proteins_g
            dish.fats_g = item.fats_g
            dish.carbs_g = item.carbs_g
            if item.calories:
                dish.calories = item.calories
            if item.weight_grams and dish.weight_grams is None:
                dish.weight_grams = item.weight_grams
            matched += 1

        await session.commit()

    await engine.dispose()
    print(f"Сопоставлено: {matched} (из них по похожести: {fuzzy}), без данных: {missed}")


if __name__ == "__main__":
    asyncio.run(run())
